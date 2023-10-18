# -*- coding: utf-8 -*-
# Part of Odoo. See LICENSE file for full copyright and licensing details.
import base64
import threading
from odoo import _, models, api, fields
from odoo.exceptions import ValidationError
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from datetime import date, datetime
from dateutil.relativedelta import relativedelta


class HrPayslip(models.Model):
    _inherit = 'hr.payslip'

    tax_regime = fields.Selection([('old_regime', 'Old'), ('new_regime', 'New')], string='Regime')
    grand_total = fields.Float(string="Total", default=0)
    is_blocked = fields.Boolean()
    full_final_emp_payslip_id = fields.Many2one('employee.full.final', string="Full & Final Employee Payslip")

    def action_payslip_done(self):
        """Checking auto email option is set.if set email containing payslip
        details will be send on confirmation"""
        res = super(HrPayslip, self).action_payslip_done()
        if self.env['ir.config_parameter'].sudo().get_param(
                'send_payslip_by_email'):
            for payslip in self:
                if payslip.employee_id.work_email:
                    template = self.env.ref(
                        'aspl_indian_payroll.email_template_payslip')
                    template.sudo().send_mail(payslip.id, force_send=True)
        return res

    def action_payslip_send(self):
        """opens a window to compose an email,
        with template message loaded by default"""
        self.ensure_one()
        ir_model_data = self.env['ir.model.data']
        try:
            template_id = \
                ir_model_data._xmlid_lookup('aspl_indian_payroll.email_template_payslip')[2]
        except ValueError:
            template_id = False
        try:
            compose_form_id = ir_model_data._xmlid_lookup('mail.email_compose_message_wizard_form')[2]
        except ValueError:
            compose_form_id = False
        ctx = {
            'default_model': 'hr.payslip',
            'default_res_id': self.ids[0],
            'default_use_template': bool(template_id),
            'default_template_id': template_id,
            'default_composition_mode': 'comment',
        }
        return {
            'name': _('Compose Email'),
            'type': 'ir.actions.act_window',
            'view_mode': 'form',
            'res_model': 'mail.compose.message',
            'views': [(compose_form_id, 'form')],
            'view_id': compose_form_id,
            'target': 'new',
            'context': ctx,
        }

    @api.model
    def get_contract(self, employee, date_from, date_to):

        # a contract is valid if it ends between the given dates
        clause_1 = ['&', ('date_end', '<=', date_to), ('date_end', '>=', date_from)]
        # OR if it starts between the given dates
        clause_2 = ['&', ('date_start', '<=', date_to), ('date_start', '>=', date_from)]
        # OR if it starts before the date_from and finish after the date_end (or never finish)
        clause_3 = ['&', ('date_start', '<=', date_from), '|', ('date_end', '=', False), ('date_end', '>=', date_to)]
        clause_final = [('employee_id', '=', employee.id), '|',
                        '|'] + clause_1 + clause_2 + clause_3

        return self.env['hr.contract'].search(clause_final).ids

    @api.model
    def create(self, vals):

        # checking it_declaration in locked stage.
        current_date = date.today()
        it_declaration_year = str(current_date.strftime('%Y')) + '-' + str(
            (current_date + relativedelta(years=1)).strftime('%y')) if current_date > date(current_date.year, 3,
                                                                                           31) else str(
            (current_date - relativedelta(years=1)).strftime('%Y')) + '-' + str(current_date.strftime('%y'))
        financial_year = self.env['financial.year'].search([('name', '=', it_declaration_year)])

        if financial_year:
            it_declaration_year_record = self.env['it.declaration.payslip'].search(
                [('employee_id', '=', vals['employee_id']), ('financial_year', '=', financial_year.id)], limit=1,
                order='create_date desc')
            employee_id = self.env['hr.employee'].search([('id', '=', vals['employee_id'])])
            if it_declaration_year_record and it_declaration_year_record.status == 'unlocked':
                raise ValidationError(
                    "IT Declaration for following employee is not locked. Please lock before you generate payslip.\n- " + str(
                        employee_id.name))
        else:
            raise ValidationError("It declaration are not created for current Year. Please create for all employee.")

        # _logger.info("Payslip parameters: %s",vals)
        payslip = super(HrPayslip, self).create(vals)
        lop_to_add = True
        payslip_obj = self.env['hr.payslip'].search([('id', '=', payslip.id)])
        for worked_line in payslip_obj.worked_days_line_ids:
            worked_line_obj = self.env['hr.payslip.worked_days'].search([('id', '=', worked_line.id)])
            lop_to_add = False if worked_line_obj.code == 'LOP' else True

        if lop_to_add:
            payslip.write({'worked_days_line_ids': [(0, 0, {'name': 'Leaves without pay', 'sequence': 2, 'code': 'LOP',
                                                            'number_of_days': 0, 'number_of_hours': 0,
                                                            'contract_id': payslip.contract_id.id})]})

        # Creating Payslips In Full & Final Employee
        full_final_emp_obj = self.env['employee.full.final'].sudo().search([('employee_id', '=', employee_id.id)])
        if full_final_emp_obj:
            full_final_emp_payslips = self.env['hr.payslip'].sudo().search(
                [('employee_id', '=', employee_id.id), ('date_from', '>=', full_final_emp_obj.resign_date)])
            full_final_emp_obj.payslip_line_ids = [(6, 0, full_final_emp_payslips.ids)]

        return payslip

    def write(self, vals):
        payslip = super(HrPayslip, self).write(vals)
        for obj in self:
            lop_to_add = True
            payslip_obj = self.env['hr.payslip'].search([('id', '=', obj.id)])
            for worked_line in payslip_obj.worked_days_line_ids:
                worked_line_obj = self.env['hr.payslip.worked_days'].search([('id', '=', worked_line.id)])
                lop_to_add = False if worked_line_obj.code == 'LOP' else True

            if lop_to_add:
                payslip_obj.write({'worked_days_line_ids': [(0, 0, {'name': 'Leaves without pay', 'sequence': 2, 'code':'LOP', 'number_of_days': 0,
                               'number_of_hours': 0,
                               'contract_id': payslip_obj.contract_id.id})]})
        return payslip

    def compute_sheet(self):
        for self in self:
            payslip = super(HrPayslip, self).compute_sheet()
            current_contract = self.env['hr.contract'].search(
                [('employee_id', '=', self.employee_id.id), ('date_start', '<', date.today()),
                 ('date_end', '>', date.today())])
            for details in current_contract.bonus_ids:
                details.paid_date = date.today()
            for details in current_contract.compensation_ids:
                details.paid_date = date.today()

    def get_projected_taxable_income_all(self):
        payslip_components = {}
        for payslip in self:
            lines = self._get_payslip_lines(payslip.contract_id.ids, payslip.id)
            for line in lines:
                code = line.get('code')
                amount = line.get('amount')
                if code in payslip_components:
                    payslip_components[code] = payslip_components[code] + amount
                else:
                    payslip_components[code] = amount
        return payslip_components


class HrPayslipRun(models.Model):
    _inherit = 'hr.payslip.run'

    company_id = fields.Many2one('res.company', 'Company')

    def _process_employee_payslip_queue(self):
        if self.env.context.get("from_backend"):
            new_cr = self.pool.cursor()
            self = self.with_env(self.env(cr=new_cr))
            if 'record_payslip_batch' in self.env.context and self.env.context['record_payslip_batch']:
                payslip_run = self.env['hr.payslip.run'].sudo().search(
                    [('id', '=', int(self.env.context['record_payslip_batch']))])
                if payslip_run.slip_ids:
                    template_id = self.env.ref('aspl_indian_payroll.employee_payslip_mail_template')
                    for payslip in payslip_run.slip_ids:
                        context = {
                            'mail_to': payslip.employee_id.work_email,
                        }
                        template_id.with_context(context).send_mail(payslip.id, force_send=True)
                        self._cr.commit()
            new_cr.close()

    def close_payslip_run(self):
        close_call = super(HrPayslipRun, self).close_payslip_run()
        if self.slip_ids:
            cron = self.env.ref('aspl_indian_payroll.ir_cron_employee_payslip_queue')
            context = {
                'record_payslip_batch': self.id,
                'from_backend': True
            }
            threaded_calculation = threading.Thread(target=self.with_context(context)._process_employee_payslip_queue,
                                                    args=())
            threaded_calculation.start()
        return close_call

    def compute_all(self):
        for slip in self.slip_ids:
            slip.compute_sheet()

    def auto_genarate_payslip1(self):
        resign_id = self.env['employee.full.final'].search(
            [('resign_date', '>=', self.date_start), ('resign_date', '<=', self.date_end),
             ('company_id', '=', self.company_id.id), ('state', '=', 'draft')])
        last_id = self.env['employee.full.final'].search(
            [('last_date', '>=', self.date_start), ('last_date', '<=', self.date_end),
             ('company_id', '=', self.company_id.id), ('state', '=', 'draft')])
        leave_id = self.env['hr.leave'].search(
            [('state', '=', 'confirm'), ('date_from', '>=', self.date_start), ('date_to', '<=', self.date_end),
             ('employee_company_id', '=', self.company_id.id)])
        contract_id = self.env['hr.contract'].search(
            [('date_start', '>=', self.date_start), ('date_start', '<=', self.date_end)])
        result_id = self.env['check.attendance.shortfall'].create(
            {'full_final_ids': last_id.ids, 'contract_ids': contract_id.ids, 'leave_ids': leave_id.ids,
             'employee_ids': resign_id.ids, 'payslip_run_id': self.id})
        return {
            'name': _('Validation Wizard'),
            'type': 'ir.actions.act_window',
            'view_mode': 'form',
            'res_model': 'check.attendance.shortfall',
            'res_id': result_id.id,
            'target': 'new'
        }

    def generate_payslips(self, vals):
        payslip_ = super(HrPayslipRun, self).generate_payslips()
        payslips = self.env['hr.payslip'].search([('state', '=', 'draft')])
        res = {
            'state': 'done'
        }
        payslips += self.env['hr.payslip'].write(res)
        return payslip_

    def bank_sheet(self):
        view_id = self.env.ref('aspl_indian_payroll.payslip_file_wizard')
        return {
            'type': 'ir.actions.act_window',
            'res_model': 'payslip.file.wizard',
            'view_type': 'form',
            'view_mode': 'form',
            'name': _('Generate File'),
            'target': 'new',
            'view_id': view_id.id,
            'context': {'batch_id': self.id},
        }

    def _get_template(self):
        self.statement_template = base64.b64encode(open("/tmp/Salary_Statement.xlsx", "rb").read())

    statement_template = fields.Binary('Template', compute="_get_template", default=date.today())

    def get_contract_template(self):
        salary_statement_details = self.env['hr.payslip.run'].search([])
        return {
            'type': 'ir.actions.act_url',
            'name': 'contract',
            'url': '/web/content/hr.payslip.run/%s/statement_template/Salary_Statement.xlsx?download=true' % (self.id),
        }

    def generate_excel_salary_statement(self, **post):
        wb = Workbook()
        ws = wb.active

        ws.cell(row=1, column=2).value = "Name"
        ws.cell(row=1, column=2).fill = PatternFill(start_color="999999", fill_type="solid")
        ws.cell(row=1, column=3).value = "Total Working Days"
        ws.cell(row=1, column=3).fill = PatternFill(start_color="999999", fill_type="solid")
        ws.cell(row=1, column=5).value = "Effective Worked Days"
        ws.cell(row=1, column=5).fill = PatternFill(start_color="999999", fill_type="solid")

        rowcounter = 2
        rule_positions = {}
        position_counter = 6  # salary according to salary_rules
        positioncounter = 4  # count of days:LOP,SHORTFALL,GLOBAL
        row_counter = 2

        master_set = set()
        for record in self.slip_ids.line_ids:
            record_value = record.salary_rule_id.sequence
            master_set.add(record_value)
        master_set_sorted = sorted(master_set)
        master_set_sorted_list = list(master_set_sorted)

        # for assigning header as per the sequence
        for rule in master_set_sorted_list:
            id_acc_to_seq = self.env['hr.salary.rule'].search([('sequence', '=', rule)])
            rule_positions[id_acc_to_seq.code] = position_counter
            ws.cell(row=1, column=position_counter).fill = PatternFill(start_color="999999", fill_type="solid")
            ws.cell(row=1, column=position_counter).value = id_acc_to_seq.name
            position_counter += 1

        # counting number of days
        for data in self.slip_ids:
            # number of working days , after subracting public holidays and weekends
            leaves = self.env['resource.calendar.leaves'].search_count(
                [('date_from', '>=', self.date_start.strftime("%Y/%m/%d, 00:00:00")),
                 ('date_to', '<=', self.date_end.strftime("%Y/%m/%d, 23:59:59")),
                 ('company_id', '=', self.company_id.id)
                    , ('resource_id', '=', False)])
            total_days = data.employee_id.get_work_days_data(datetime.combine(self.date_start, datetime.min.time()),
                                                             datetime.combine(self.date_end, datetime.max.time()),
                                                             compute_leaves=False)
            # not_worked = float()
            leave_lop = 0
            leave_sf = 0
            leave_sf_2 = 0
            for normal_work_days in data.worked_days_line_ids:
                if normal_work_days.code == "LOP":
                    leave_lop = normal_work_days.number_of_days
                    # print(leave_lop,"LOP leaves")
                if normal_work_days.code == "SHORTFALL":
                    leave_sf = normal_work_days.number_of_days
                    leave_sf_hrs = normal_work_days.number_of_hours
                    if leave_sf_hrs > 2 and leave_sf_hrs < 4:
                        leave_sf_2 = 0.5
                    elif leave_sf_hrs > 3 and leave_sf_hrs < 9:
                        leave_sf_2 = 1
                not_worked = leave_lop + leave_sf + leaves + leave_sf_2
            ws.cell(row=rowcounter, column=2).value = data.employee_id.name
            ws.cell(row=rowcounter, column=3).value = total_days['days'] - leaves
            ws.cell(row=rowcounter, column=5).value = total_days['days'] - not_worked

            for work_info in data.worked_days_line_ids:
                if work_info.code in rule_positions:
                    ws.cell(row=row_counter, column=rule_positions.get(work_info.code)).value = work_info.number_of_days

                else:
                    if work_info.code in ['LOP']:  # ,'GLOBAL','SHORTFALL'
                        rule_positions[work_info.code] = positioncounter
                        ws.cell(row=1, column=positioncounter).fill = PatternFill(start_color="999999",
                                                                                  fill_type="solid")
                        ws.cell(row=1, column=positioncounter).value = work_info.name
                        ws.cell(row=row_counter, column=positioncounter).value = work_info.number_of_days
                        positioncounter += 1
            row_counter += 1

            # main salary according to rules
            for info in data.line_ids:
                # if header code is already present
                if info.salary_rule_id.code in rule_positions:
                    ws.cell(row=rowcounter, column=rule_positions[info.salary_rule_id.code]).value = info.total
                # If code is not present
                else:
                    rule_positions[info.salary_rule_id.code] = position_counter
                    ws.cell(row=1, column=position_counter).fill = PatternFill(start_color="999999", fill_type="solid")
                    ws.cell(row=1, column=position_counter).value = info.salary_rule_id.name
                    ws.cell(row=rowcounter, column=position_counter).value = info.total
                    position_counter += 1

            rowcounter += 1

        wb.save("/tmp/Salary_Statement.xlsx")

        return self.get_contract_template()


class HrPayslipEmployees(models.TransientModel):
    _inherit = 'hr.payslip.employees'

    def _current_working_employee(self):
        if 'active_id' in self.env.context and 'active_model' in self.env.context and self.env.context.get(
                'active_model') == 'hr.payslip.run':
            current_payslip_batch = self.env[self.env.context.get('active_model')].search(
                [("id", '=', self.env.context.get('active_id'))])
            employee = self.env['hr.contract'].search(
                [('date_end', '>=', current_payslip_batch.date_start), ('state', '=', 'open')]).employee_id

            if 'allowed_company_ids' in self.env.context:
                current_company_employee = employee.filtered(
                    lambda emp: emp.company_id.id in self.env.context.get('allowed_company_ids'))
                return [('id', 'in', current_company_employee.ids)]
            else:
                return [('id', 'in', employee.ids)]
        else:
            return [('id', 'in', [])]

    employee_ids = fields.Many2many('hr.employee', 'hr_employee_group_rel', 'payslip_id', 'employee_id', 'Employees',
                                    domain=_current_working_employee)

    def compute_sheet(self):
        if self.employee_ids:
            current_date = date.today()
            employee_list = []
            it_declaration_year = str(current_date.strftime('%Y')) + '-' + str(
                (current_date + relativedelta(years=1)).strftime('%y')) if current_date > date(current_date.year, 3,
                                                                                               31) else str(
                (current_date - relativedelta(years=1)).strftime('%Y')) + '-' + str(current_date.strftime('%y'))
            financial_year = self.env['financial.year'].search([('name', '=', it_declaration_year)])
            if financial_year:
                for employee_id in self.employee_ids:
                    it_declaration_year_record = self.env['it.declaration.payslip'].search(
                        [('employee_id', '=', employee_id.id), ('financial_year', '=', financial_year.id)], limit=1,
                        order='create_date desc')
                    if it_declaration_year_record and it_declaration_year_record.status == 'unlocked':
                        employee_list.append(employee_id.name)
            else:
                raise ValidationError(
                    "It declaration are not created for current Year. Please create for all employee.")

            if len(employee_list) > 1:
                raise ValidationError(
                    "IT Declaration for following employees are not locked. Please lock before you generate payslips.\n" + "\n".join(
                        "- " + employee for employee in employee_list))
            elif len(employee_list) == 1:
                raise ValidationError(
                    "IT Declaration for following employee is not locked. Please lock before you generate payslip.\n- " + str(
                        employee_list[0]))
        if self.self.employee_ids.emp_state != 'on_notice':
            super(HrPayslipEmployees, self).compute_sheet()
